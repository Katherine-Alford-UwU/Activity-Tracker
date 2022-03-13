from openpyxl import Workbook, load_workbook
from Tracker import StepTracker, SleepTracker, WeatherTracker

workbook = load_workbook("trackerdata.xlsx")
sheet = workbook.get_sheet_by_name('Sheet1')

track = WeatherTracker("Katherine",12)
track2 = SleepTracker("Katherine",12)
track3 = StepTracker("Katherine", 12)

date1 = input("What is the date? MM/DD/YY")
watHrs = input("How many hours did you sleep last night?")
watSteps = int(input("How many steps did you take today?"))
watWethur = input("What was the high and low of today's weather? low/high")

workbookRead = load_workbook ('trackerdata.xlsx')
sheetName = workbookRead['Sheet1']

emptyRow = 0
for i in range (1,1000000):
    if (sheetName['A' + str(i)].value == None):
        emptyRow = i
        break

sheet.cell(column=1, row=emptyRow, value=date1)
sheet.cell(column=2,row=emptyRow, value=watSteps)
sheet.cell(column=3,row=emptyRow, value=watHrs)
sheet.cell(column=4,row=emptyRow, value=watWethur)

workbook.save(filename="trackerdata.xlsx")
